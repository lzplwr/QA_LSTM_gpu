# -*- coding: utf-8 -*-
"""
Created on Thu May 17 17:59:40 2018

@author: hasee
"""

import modules.util as util
from openpyxl import Workbook
import random
from pickle import dump
import os

negative_pool = 100

code_path = os.path.realpath(__file__)
dir_path = os.path.dirname(code_path)

# 获取新的训练集、验证集和测试集

data = util.read_table('corpus.xlsx')
random.shuffle(data)

test_data = data[-2000:]
train_data = data[:-2000]
dev_data = train_data[-1000:]
train_data = train_data[:-1000]

wb = Workbook()
sheet = wb.active
for i,(q,a) in enumerate(train_data):
    sheet.cell(row = i+1, column = 1).value = q
    sheet.cell(row = i+1, column = 2).value = a

file_path = os.path.join(dir_path, 'data', 'train.xlsx')
wb.save(file_path)

wb = Workbook()
sheet = wb.active
for i,(q,a) in enumerate(dev_data):
    sheet.cell(row = i+1, column = 1).value = q
    sheet.cell(row = i+1, column = 2).value = a
    negative_answers = random.sample(range(len(data)), negative_pool)
    negative_answers = [ str(k) for k in negative_answers ]
    negative_answers = ' '.join(negative_answers)
    sheet.cell(row = i+1, column = 3).value = negative_answers

file_path = os.path.join(dir_path, 'data', 'dev.xlsx')
wb.save(file_path)

wb = Workbook()
sheet = wb.active
for i,(q,a) in enumerate(test_data):
    sheet.cell(row = i+1, column = 1).value = q
    sheet.cell(row = i+1, column = 2).value = a
    negative_answers = random.sample(range(len(data)), negative_pool)
    negative_answers = [ str(k) for k in negative_answers ]
    negative_answers = ' '.join(negative_answers)
    sheet.cell(row = i+1, column = 3).value = negative_answers

file_path = os.path.join(dir_path, 'data', 'test.xlsx')
wb.save(file_path)

# 获取关键词和它们的idf值

idf_dict, gram2ques = util.keyword_cal()

file_path = os.path.join(dir_path, 'model', 'gram2ques')
with open(file_path, 'wb') as f:
    dump(gram2ques, f)

file_path = os.path.join(dir_path, 'model', 'idf_dict')
with open(file_path, 'wb') as f:
    dump(idf_dict, f)